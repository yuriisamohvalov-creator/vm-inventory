"""Экспорт отчёта в XLSX формат."""
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .models import Department, VM


def build_report_xlsx():
    """Build hierarchical report XLSX (Department -> Stream -> IS -> VMs)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "VM Inventory Report"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    
    # Заголовок
    ws.merge_cells('A1:M1')
    ws['A1'] = 'VM Inventory — Report'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    row = 3
    
    # Стили
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    dept_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    dept_font = Font(bold=True)
    stream_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    stream_font = Font(bold=True)
    is_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    
    departments = Department.objects.prefetch_related(
        'streams__info_systems__vms'
    ).all()
    orphan = VM.objects.filter(info_system__isnull=True)
    
    for dept in departments:
        dept_vm_count = 0
        dept_sum_cpu = 0
        dept_sum_ram = 0
        dept_sum_disk = 0
        
        # Департамент
        dept_has_exceeded = False
        dept_name = dept.name
        ws.merge_cells(f'A{row}:M{row}')
        ws[f'A{row}'] = dept_name
        ws[f'A{row}'].fill = dept_fill
        ws[f'A{row}'].font = dept_font
        row += 1
        
        for stream in dept.streams.all():
            stream_vm_count = 0
            stream_sum_cpu = 0
            stream_sum_ram = 0
            stream_sum_disk = 0
            
            # Стрим
            ws.merge_cells(f'B{row}:M{row}')
            ws[f'B{row}'] = stream.name
            ws[f'B{row}'].fill = stream_fill
            ws[f'B{row}'].font = stream_font
            row += 1
            
            for isys in stream.info_systems.all():
                vms_list = list(isys.vms.all())
                if vms_list:
                    # ИС заголовок
                    ws.merge_cells(f'C{row}:M{row}')
                    ws[f'C{row}'] = isys.name
                    ws[f'C{row}'].fill = is_fill
                    ws[f'C{row}'].font = Font(bold=True)
                    row += 1
                    
                    # Заголовки колонок ВМ
                    ws[f'D{row}'] = 'FQDN'
                    ws[f'E{row}'] = 'IP'
                    ws[f'F{row}'] = 'CPU'
                    ws[f'G{row}'] = 'RAM (ГБ)'
                    ws[f'H{row}'] = 'Диск (ГБ)'
                    ws[f'I{row}'] = 'БА.ПФМ_зак'
                    ws[f'J{row}'] = 'БА.ПФМ_исп'
                    ws[f'K{row}'] = 'БА.Программа_бюджета'
                    ws[f'L{row}'] = 'БА.Финансовая_позиция'
                    ws[f'M{row}'] = 'БА.Mir-код'
                    for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
                        ws[f'{col}{row}'].fill = header_fill
                        ws[f'{col}{row}'].font = header_font
                    row += 1
                    
                    # ВМ
                    for vm in vms_list:
                        vm_fqdn = vm.fqdn
                        if vm.info_system is None:
                            vm_fqdn += " [ИС УДАЛЕНА]"
                        ws[f'D{row}'] = vm_fqdn
                        ws[f'E{row}'] = vm.ip
                        ws[f'F{row}'] = vm.cpu
                        ws[f'G{row}'] = vm.ram
                        ws[f'H{row}'] = vm.disk
                        ws[f'I{row}'] = vm.ba_pfm_zak
                        ws[f'J{row}'] = vm.ba_pfm_isp
                        ws[f'K{row}'] = vm.ba_programma_byudzheta
                        ws[f'L{row}'] = vm.ba_finansovaya_pozitsiya
                        ws[f'M{row}'] = vm.ba_mir_kod
                        row += 1
                    
                    # Итого ИС
                    sum_cpu = sum(vm.cpu for vm in vms_list)
                    sum_ram = sum(vm.ram for vm in vms_list)
                    sum_disk = sum(vm.disk for vm in vms_list)
                    ws[f'D{row}'] = f'Итого ИС: {len(vms_list)} ВМ'
                    ws[f'F{row}'] = sum_cpu
                    ws[f'G{row}'] = sum_ram
                    ws[f'H{row}'] = sum_disk
                    ws[f'D{row}'].font = Font(bold=True)
                    row += 1
                    
                    stream_vm_count += len(vms_list)
                    stream_sum_cpu += sum_cpu
                    stream_sum_ram += sum_ram
                    stream_sum_disk += sum_disk
                
            if stream_vm_count > 0:
                # Итого Стрим (одна строка на стрим, после всех ИС)
                stream_has_exceeded = (
                    (stream.cpu_quota > 0 and stream_sum_cpu > stream.cpu_quota)
                    or (stream.ram_quota > 0 and stream_sum_ram > stream.ram_quota)
                    or (stream.disk_quota > 0 and stream_sum_disk > stream.disk_quota)
                )
                stream_tot = f'Итого Стрим: {stream_vm_count} ВМ'
                if stream.cpu_quota > 0:
                    stream_tot += f', CPU: {stream_sum_cpu}/{stream.cpu_quota}'
                else:
                    stream_tot += f', CPU: {stream_sum_cpu}'
                if stream.ram_quota > 0:
                    stream_tot += f', RAM: {stream_sum_ram}/{stream.ram_quota} ГБ'
                else:
                    stream_tot += f', RAM: {stream_sum_ram} ГБ'
                if stream.disk_quota > 0:
                    stream_tot += f', Диск: {stream_sum_disk}/{stream.disk_quota} ГБ'
                else:
                    stream_tot += f', Диск: {stream_sum_disk} ГБ'
                ws[f'C{row}'] = f'{"🚨 " if stream_has_exceeded else ""}{stream_tot}'
                ws[f'F{row}'] = stream_sum_cpu
                ws[f'G{row}'] = stream_sum_ram
                ws[f'H{row}'] = stream_sum_disk
                ws[f'C{row}'].font = Font(bold=True)
                ws[f'C{row}'].fill = stream_fill
                row += 1
                
                dept_vm_count += stream_vm_count
                dept_sum_cpu += stream_sum_cpu
                dept_sum_ram += stream_sum_ram
                dept_sum_disk += stream_sum_disk
        
        if dept_vm_count > 0:
            # Проверка превышения квот
            if dept.cpu_quota > 0 and dept_sum_cpu > dept.cpu_quota:
                dept_has_exceeded = True
            if dept.ram_quota > 0 and dept_sum_ram > dept.ram_quota:
                dept_has_exceeded = True
            if dept.disk_quota > 0 and dept_sum_disk > dept.disk_quota:
                dept_has_exceeded = True
            
            # Итого Департамент
            dept_tot = f'{"🚨 " if dept_has_exceeded else ""}Итого Департамент: {dept_vm_count} ВМ'
            if dept.cpu_quota > 0:
                dept_tot += f', CPU: {dept_sum_cpu}/{dept.cpu_quota}'
            else:
                dept_tot += f', CPU: {dept_sum_cpu}'
            if dept.ram_quota > 0:
                dept_tot += f', RAM: {dept_sum_ram}/{dept.ram_quota} ГБ'
            else:
                dept_tot += f', RAM: {dept_sum_ram} ГБ'
            if dept.disk_quota > 0:
                dept_tot += f', Диск: {dept_sum_disk}/{dept.disk_quota} ГБ'
            else:
                dept_tot += f', Диск: {dept_sum_disk} ГБ'
            
            ws[f'A{row}'] = dept_tot
            ws[f'F{row}'] = dept_sum_cpu
            ws[f'G{row}'] = dept_sum_ram
            ws[f'H{row}'] = dept_sum_disk
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = dept_fill
            row += 1
    
    # ВМ без ИС
    if orphan.exists():
        orphan_list = list(orphan)
        ws.merge_cells(f'A{row}:M{row}')
        ws[f'A{row}'] = '(ВМ без ИС / удалённая ИС)'
        ws[f'A{row}'].fill = dept_fill
        ws[f'A{row}'].font = dept_font
        row += 1
        
        ws[f'D{row}'] = 'FQDN'
        ws[f'E{row}'] = 'IP'
        ws[f'F{row}'] = 'CPU'
        ws[f'G{row}'] = 'RAM (ГБ)'
        ws[f'H{row}'] = 'Диск (ГБ)'
        ws[f'I{row}'] = 'БА.ПФМ_зак'
        ws[f'J{row}'] = 'БА.ПФМ_исп'
        ws[f'K{row}'] = 'БА.Программа_бюджета'
        ws[f'L{row}'] = 'БА.Финансовая_позиция'
        ws[f'M{row}'] = 'БА.Mir-код'
        for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
            ws[f'{col}{row}'].fill = header_fill
            ws[f'{col}{row}'].font = header_font
        row += 1
        
        for vm in orphan_list:
            ws[f'D{row}'] = vm.fqdn + " [ИС УДАЛЕНА]"
            ws[f'E{row}'] = vm.ip
            ws[f'F{row}'] = vm.cpu
            ws[f'G{row}'] = vm.ram
            ws[f'H{row}'] = vm.disk
            ws[f'I{row}'] = vm.ba_pfm_zak
            ws[f'J{row}'] = vm.ba_pfm_isp
            ws[f'K{row}'] = vm.ba_programma_byudzheta
            ws[f'L{row}'] = vm.ba_finansovaya_pozitsiya
            ws[f'M{row}'] = vm.ba_mir_kod
            row += 1
        
        sum_cpu = sum(vm.cpu for vm in orphan_list)
        sum_ram = sum(vm.ram for vm in orphan_list)
        sum_disk = sum(vm.disk for vm in orphan_list)
        ws[f'D{row}'] = f'Итого: {len(orphan_list)} ВМ'
        ws[f'F{row}'] = sum_cpu
        ws[f'G{row}'] = sum_ram
        ws[f'H{row}'] = sum_disk
        ws[f'D{row}'].font = Font(bold=True)
    
    # Автоподбор ширины колонок
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
